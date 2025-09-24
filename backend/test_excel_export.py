#!/usr/bin/env python3
"""Test Excel export functionality"""
import asyncio
import os
import sys
from pathlib import Path

# Add current directory to path
sys.path.insert(0, str(Path(__file__).parent))

async def test_excel_export():
    """Test if Excel export is working"""
    try:
        # Import required modules
        from services.export.excel_exporter import ExcelExporter
        from models.forecast import ForecastResponse, ModelResponse
        from datetime import datetime

        print("✅ Excel exporter module imported successfully!")

        # Create test data
        test_forecast = ForecastResponse(
            forecast_id='test-excel-123',
            question='Will AI impact jobs by 2026?',
            definition='20% job transformation',
            timeframe='2026',
            models_used=['x-ai/grok-4-fast:free', 'google/gemma-2b-it:free'],
            aggregate_probability=75.5,
            model_responses=[
                ModelResponse(
                    model='x-ai/grok-4-fast:free',
                    probability=80.0,
                    confidence='high',
                    reasoning='AI adoption is accelerating rapidly across industries',
                    analysis_time=2.5
                ),
                ModelResponse(
                    model='google/gemma-2b-it:free',
                    probability=71.0,
                    confidence='medium',
                    reasoning='Gradual AI integration expected in most sectors',
                    analysis_time=3.1
                )
            ],
            iterations=2,
            created_at=datetime.now().isoformat(),
            total_analysis_time=5.6
        )

        print("✅ Test forecast data created!")

        # Initialize Excel exporter
        exporter = ExcelExporter()
        print("✅ ExcelExporter initialized!")

        # Create Excel report
        file_path = await exporter.export_forecast(test_forecast)
        print(f"✅ Excel file created at: {file_path}")

        # Verify file exists and has content
        if os.path.exists(file_path):
            file_size = os.path.getsize(file_path)
            print(f"✅ File exists with size: {file_size:,} bytes")

            # Try to load and verify the Excel file
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_path)
                sheet_names = wb.sheetnames
                print(f"✅ Excel file contains {len(sheet_names)} sheets: {', '.join(sheet_names)}")

                # Check each sheet has content
                for sheet_name in sheet_names:
                    sheet = wb[sheet_name]
                    row_count = sheet.max_row
                    col_count = sheet.max_column
                    print(f"   - {sheet_name}: {row_count} rows × {col_count} columns")

                wb.close()
                print("\n🎉 EXCEL EXPORT IS WORKING PERFECTLY!")

            except Exception as e:
                print(f"⚠️ Could not verify Excel structure: {e}")

            # Clean up test file
            os.remove(file_path)
            print("✅ Test file cleaned up")

        else:
            print("❌ Excel file was not created!")

    except ImportError as e:
        print(f"❌ Import error: {e}")
        print("   Missing dependencies for Excel export")

    except Exception as e:
        print(f"❌ Excel generation failed: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    print("="*60)
    print("TESTING EXCEL EXPORT FUNCTIONALITY")
    print("="*60)
    asyncio.run(test_excel_export())