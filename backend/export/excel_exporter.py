"""Enhanced Excel export functionality with professional visualizations"""
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle,
    GradientFill, Protection
)
from openpyxl.chart import (
    BarChart, Reference, LineChart, ScatterChart, PieChart,
    AreaChart, Series
)
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.text import RichText
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import SolidColorFillProperties, ColorChoice
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO

from core.models import ForecastResult, ExportData
from config.settings import get_settings
from utils.visual import visual

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Enhanced Excel exporter with professional visualizations and styling"""

    def __init__(self):
        self.settings = get_settings()
        self._setup_styles()

    def _setup_styles(self):
        """Setup professional Excel styles"""
        # Header style
        self.header_style = NamedStyle(name="header_style")
        self.header_style.font = Font(bold=True, color="FFFFFF", size=12)
        self.header_style.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_style.alignment = Alignment(horizontal="center", vertical="center")
        self.header_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Title style
        self.title_style = NamedStyle(name="title_style")
        self.title_style.font = Font(bold=True, size=16, color="2E5C8A")
        self.title_style.alignment = Alignment(horizontal="center", vertical="center")

        # Data style
        self.data_style = NamedStyle(name="data_style")
        self.data_style.alignment = Alignment(horizontal="center", vertical="center")
        self.data_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Highlight style for key metrics
        self.highlight_style = NamedStyle(name="highlight_style")
        self.highlight_style.font = Font(bold=True, size=14, color="FFFFFF")
        self.highlight_style.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        self.highlight_style.alignment = Alignment(horizontal="center", vertical="center")

    def export_forecast(
        self,
        forecast_result: ForecastResult,
        filename: Optional[str] = None,
        output_dir: Optional[Path] = None
    ) -> Path:
        """
        Export complete forecast results to Excel

        Args:
            forecast_result: The forecast results to export
            filename: Custom filename (auto-generated if None)
            output_dir: Output directory (uses config default if None)

        Returns:
            Path to the created Excel file
        """
        # Generate filename if not provided
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.settings.output.excel_filename_prefix}_{timestamp}.xlsx"

        # Use configured output directory if not provided
        if output_dir is None:
            output_dir = self.settings.output.output_dir

        output_path = output_dir / filename

        # Create export data
        export_data = ExportData(forecast_result=forecast_result)

        # Create workbook with enhanced styling
        workbook = Workbook()

        # Add custom styles to workbook
        workbook.add_named_style(self.header_style)
        workbook.add_named_style(self.title_style)
        workbook.add_named_style(self.data_style)
        workbook.add_named_style(self.highlight_style)

        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])

        # Generate visualizations first
        chart_dir = output_dir / "charts"
        chart_dir.mkdir(exist_ok=True)

        self._generate_visualizations(export_data, chart_dir)

        # Create enhanced sheets (simplified for compatibility)
        self._create_enhanced_summary_sheet(workbook, export_data, chart_dir)
        self._create_enhanced_responses_sheet(workbook, export_data)
        self._create_enhanced_statistical_sheet(workbook, export_data, chart_dir)
        self._create_advanced_algorithms_sheet(workbook, export_data)
        # Skip visual analysis sheet for now due to merged cell issue
        self._create_reasoning_analysis_sheet(workbook, export_data)
        self._create_metadata_sheet(workbook, export_data)

        # Add new enhanced features
        self._create_probability_heatmap_sheet(workbook, export_data)
        self._create_executive_brief_sheet(workbook, export_data)

        # Save workbook
        workbook.save(output_path)
        logger.info(f"Enhanced Excel export saved to: {output_path}")

        return output_path

    def _generate_visualizations(self, export_data: ExportData, chart_dir: Path):
        """Generate matplotlib visualizations for embedding"""
        forecast_result = export_data.forecast_result

        # Model comparison chart
        visual.create_model_comparison_chart(
            forecast_result.statistics.model_stats,
            chart_dir
        )

        # Probability distribution
        visual.create_probability_distribution_chart(
            forecast_result.responses,
            chart_dir
        )

        # Timeline chart
        visual.create_timeline_chart(
            forecast_result.responses,
            chart_dir
        )

    def _create_enhanced_summary_sheet(self, workbook: Workbook, export_data: ExportData, chart_dir: Path):
        """Create enhanced summary dashboard with professional styling and charts"""
        ws = workbook.create_sheet("📊 Executive Summary")

        # Professional header with gradient
        ws['A1'] = "🔮 AI Foresight Analysis - Executive Summary"
        ws['A1'].style = self.title_style
        ws.merge_cells('A1:H1')
        ws.row_dimensions[1].height = 30

        # Forecast question box
        ws['A3'] = "Forecast Question"
        ws['A3'].style = self.header_style
        ws.merge_cells('A3:H3')

        ws['A4'] = export_data.forecast_result.metadata.question
        ws['A4'].font = Font(size=12, italic=True)
        ws['A4'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells('A4:H4')
        ws.row_dimensions[4].height = 40

        # Key Result - Ensemble Probability (large, prominent)
        ensemble_prob = export_data.forecast_result.get_ensemble_probability()
        ws['A6'] = "🎯 ENSEMBLE FORECAST"
        ws['A6'].style = self.header_style
        ws.merge_cells('A6:C6')

        if ensemble_prob:
            ws['A7'] = f"{ensemble_prob:.1f}%"
            ws['A7'].style = self.highlight_style
            ws['A7'].font = Font(size=32, bold=True, color="FFFFFF")
            ws.merge_cells('A7:C7')
            ws.row_dimensions[7].height = 50

            # Confidence indicator
            confidence_text = self._get_confidence_indicator(ensemble_prob)
            ws['A8'] = confidence_text
            ws['A8'].font = Font(size=12, italic=True)
            ws.merge_cells('A8:C8')

        # Key metrics grid
        ws['E6'] = "📈 KEY METRICS"
        ws['E6'].style = self.header_style
        ws.merge_cells('E6:H6')

        metrics = [
            ("Models Used", len(export_data.forecast_result.metadata.models)),
            ("Total Queries", export_data.forecast_result.metadata.total_queries),
            ("Success Rate", f"{export_data.forecast_result.statistics.success_rate * 100:.1f}%"),
            ("Standard Deviation", f"{export_data.forecast_result.statistics.std:.1f}%" if export_data.forecast_result.statistics.std else "N/A"),
            ("Analysis Duration", f"{export_data.forecast_result.metadata.duration_seconds:.1f}s"),
            ("Consensus Level", self._get_consensus_level(export_data.forecast_result.statistics))
        ]

        for i, (metric, value) in enumerate(metrics):
            row = 7 + i
            ws[f'E{row}'] = metric
            ws[f'E{row}'].font = Font(bold=True)
            ws[f'F{row}'] = value
            ws[f'F{row}'].style = self.data_style

        # Model Performance Table
        ws['A12'] = "🤖 MODEL PERFORMANCE BREAKDOWN"
        ws['A12'].style = self.header_style
        ws.merge_cells('A12:H12')

        # Table headers with icons
        headers = ['🤖 Model', '📊 Mean', '📈 Std Dev', '📍 Range', '🔢 Count', '✅ Success', '⭐ Quality']
        for i, header in enumerate(headers):
            cell = ws.cell(row=13, column=i+1)
            cell.value = header
            cell.style = self.header_style

        # Model data with conditional formatting
        for i, (model, stats) in enumerate(export_data.forecast_result.statistics.model_stats.items()):
            row = 14 + i
            model_name = model.split('/')[-1]

            # Quality score based on consistency
            quality_score = 1 / (1 + stats.std) if stats.std > 0 else 1
            quality_stars = "⭐" * min(5, int(quality_score * 5 + 0.5))

            ws[f'A{row}'] = model_name
            ws[f'B{row}'] = f"{stats.mean:.1f}%"
            ws[f'C{row}'] = f"±{stats.std:.1f}%"
            ws[f'D{row}'] = f"{stats.min:.0f}-{stats.max:.0f}%"
            ws[f'E{row}'] = stats.count
            ws[f'F{row}'] = f"{stats.success_rate * 100:.0f}%"
            ws[f'G{row}'] = quality_stars

            # Apply data styling
            for col in range(1, 8):
                ws.cell(row=row, column=col).style = self.data_style

        # Add conditional formatting for model performance
        self._add_conditional_formatting(ws, 14, 14 + len(export_data.forecast_result.statistics.model_stats))

        # Embed charts if available
        self._embed_charts(ws, chart_dir)

        # Auto-adjust column widths
        column_widths = [15, 12, 12, 15, 8, 10, 12, 20]
        for i, width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i+1)].width = width

    def _get_confidence_indicator(self, probability: float) -> str:
        """Get confidence indicator based on probability"""
        if 40 <= probability <= 60:
            return "🟡 Moderate Confidence - High Uncertainty"
        elif 20 <= probability <= 80:
            return "🟢 Good Confidence Level"
        else:
            return "🔴 High Confidence - Extreme Prediction"

    def _get_consensus_level(self, statistics) -> str:
        """Get consensus level description"""
        if not statistics.std:
            return "N/A"

        cv = statistics.std / statistics.mean if statistics.mean > 0 else 1
        if cv < 0.1:
            return "🟢 Very High"
        elif cv < 0.2:
            return "🟡 High"
        elif cv < 0.4:
            return "🟠 Moderate"
        else:
            return "🔴 Low"

    def _add_conditional_formatting(self, ws, start_row: int, end_row: int):
        """Add conditional formatting to the model performance table"""
        # Success rate color scale
        success_range = f'F{start_row}:F{end_row}'
        ws.conditional_formatting.add(success_range,
            ColorScaleRule(start_type='min', start_color='FFCCCB',
                          end_type='max', end_color='90EE90'))

        # Standard deviation (lower is better)
        std_range = f'C{start_row}:C{end_row}'
        ws.conditional_formatting.add(std_range,
            ColorScaleRule(start_type='min', start_color='90EE90',
                          end_type='max', end_color='FFCCCB'))

    def _embed_charts(self, ws, chart_dir: Path):
        """Embed matplotlib charts into the worksheet"""
        try:
            # Model comparison chart
            comparison_chart_path = chart_dir / "model_comparison.png"
            if comparison_chart_path.exists():
                img = Image(str(comparison_chart_path))
                img.width = 600
                img.height = 300
                ws.add_image(img, 'A25')

            # Distribution chart
            dist_chart_path = chart_dir / "probability_distribution.png"
            if dist_chart_path.exists():
                img = Image(str(dist_chart_path))
                img.width = 600
                img.height = 300
                ws.add_image(img, 'A45')

        except Exception as e:
            logger.warning(f"Could not embed charts: {e}")

    def _create_enhanced_responses_sheet(self, workbook: Workbook, export_data: ExportData):
        """Create enhanced individual responses sheet with filtering and formatting"""
        ws = workbook.create_sheet("🔍 Detailed Responses")

        # Title
        ws['A1'] = "🔍 Individual Model Responses - Complete Dataset"
        ws['A1'].style = self.title_style
        ws.merge_cells('A1:L1')
        ws.row_dimensions[1].height = 25

        # Convert to DataFrame
        rows = export_data.to_detail_rows()
        df = pd.DataFrame(rows)

        # Enhanced headers with descriptions
        enhanced_headers = [
            ("🆔 ID", "Unique Response ID"),
            ("🤖 Model", "AI Model Used"),
            ("🔄 Iteration", "Query Number"),
            ("🕐 Timestamp", "Response Time"),
            ("✅ Status", "Success/Error"),
            ("🎯 Probability", "Forecast %"),
            ("⏱️ Time (s)", "Response Duration"),
            ("📝 Reasoning", "Full Response Text"),
            ("❌ Error", "Error Message"),
            ("🔤 Prompt Tokens", "Input Tokens"),
            ("💬 Response Tokens", "Output Tokens"),
            ("📊 Total Tokens", "Total Token Usage")
        ]

        # Create header row with enhanced styling
        for i, (header, description) in enumerate(enhanced_headers):
            cell = ws.cell(row=2, column=i+1)
            cell.value = header
            cell.style = self.header_style
            # Note: Comment functionality would require openpyxl Comment object

        # Add data with conditional formatting
        for row_idx, row_data in enumerate(rows, start=3):
            for col_idx, (key, value) in enumerate(row_data.items(), start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.style = self.data_style

                # Special formatting for status column
                if key == "Status" and value == "success":
                    cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                elif key == "Status" and value != "success":
                    cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

                # Probability cell formatting
                if key == "Probability" and isinstance(value, (int, float)):
                    if 40 <= value <= 60:
                        cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                    elif value < 20 or value > 80:
                        cell.fill = PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid")

        # Add data bars for probability column
        prob_col = 6  # Probability column
        prob_range = f'{get_column_letter(prob_col)}3:{get_column_letter(prob_col)}{len(rows)+2}'
        ws.conditional_formatting.add(prob_range, DataBarRule(start_type='min', end_type='max',
                                                               color='4CAF50'))

        # Column widths
        column_widths = [12, 15, 8, 18, 8, 10, 8, 50, 20, 10, 12, 10]
        for i, width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i+1)].width = width

        # Freeze panes for easy navigation
        ws.freeze_panes = 'A3'

        # Add auto-filter
        ws.auto_filter.ref = f'A2:{get_column_letter(len(enhanced_headers))}{len(rows)+2}'

    def _create_enhanced_statistical_sheet(self, workbook: Workbook, export_data: ExportData, chart_dir: Path):
        """Create enhanced statistical analysis sheet with native Excel charts"""
        ws = workbook.create_sheet("📊 Statistical Analysis")

        # Title
        ws['A1'] = "📊 Advanced Statistical Analysis"
        ws['A1'].style = self.title_style
        ws.merge_cells('A1:H1')

        # Overall statistics section
        stats = export_data.forecast_result.statistics
        ws['A3'] = "📈 ENSEMBLE STATISTICS"
        ws['A3'].style = self.header_style
        ws.merge_cells('A3:D3')

        overall_stats = [
            ("Mean Probability", f"{stats.mean:.2f}%" if stats.mean else "N/A"),
            ("Median Probability", f"{stats.median:.2f}%" if stats.median else "N/A"),
            ("Standard Deviation", f"{stats.std:.2f}%" if stats.std else "N/A"),
            ("Minimum", f"{stats.min:.2f}%" if stats.min else "N/A"),
            ("Maximum", f"{stats.max:.2f}%" if stats.max else "N/A"),
            ("Range", f"{stats.max - stats.min:.2f}%" if stats.max and stats.min else "N/A"),
            ("Coefficient of Variation", f"{(stats.std/stats.mean):.3f}" if stats.std and stats.mean else "N/A")
        ]

        for i, (stat, value) in enumerate(overall_stats):
            ws[f'A{4+i}'] = stat
            ws[f'A{4+i}'].font = Font(bold=True)
            ws[f'B{4+i}'] = value
            ws[f'B{4+i}'].style = self.data_style

        # Model comparison table with charts
        ws['A13'] = "🤖 MODEL COMPARISON ANALYSIS"
        ws['A13'].style = self.header_style
        ws.merge_cells('A13:H13')

        # Prepare data for chart
        model_data = []
        for model, model_stats in stats.model_stats.items():
            model_data.append([
                model.split('/')[-1],
                model_stats.mean,
                model_stats.std,
                model_stats.min,
                model_stats.max,
                model_stats.count,
                model_stats.success_rate * 100
            ])

        # Headers
        headers = ['Model', 'Mean (%)', 'Std Dev (%)', 'Min (%)', 'Max (%)', 'Count', 'Success (%)']
        for i, header in enumerate(headers):
            cell = ws.cell(row=14, column=i+1)
            cell.value = header
            cell.style = self.header_style

        # Data
        for i, row_data in enumerate(model_data):
            for j, value in enumerate(row_data):
                cell = ws.cell(row=15+i, column=j+1)
                cell.value = value
                cell.style = self.data_style

        # Create Excel charts
        self._create_model_comparison_chart_excel(ws, 15, len(model_data))
        self._create_distribution_analysis(ws, export_data.forecast_result.responses)

    def _create_model_comparison_chart_excel(self, ws, start_row: int, num_models: int):
        """Create native Excel chart for model comparison"""
        # Mean comparison chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Model Performance Comparison"
        chart.y_axis.title = 'Probability (%)'
        chart.x_axis.title = 'Models'

        # Data for chart
        data = Reference(ws, min_col=2, min_row=start_row-1, max_row=start_row+num_models-1, max_col=2)
        cats = Reference(ws, min_col=1, min_row=start_row, max_row=start_row+num_models-1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Styling
        chart.series[0].graphicalProperties.solidFill = "4CAF50"

        ws.add_chart(chart, "J15")

        # Standard deviation chart
        std_chart = BarChart()
        std_chart.type = "col"
        std_chart.style = 11
        std_chart.title = "Model Consistency (Lower = Better)"
        std_chart.y_axis.title = 'Standard Deviation (%)'
        std_chart.x_axis.title = 'Models'

        std_data = Reference(ws, min_col=3, min_row=start_row-1, max_row=start_row+num_models-1, max_col=3)
        std_chart.add_data(std_data, titles_from_data=True)
        std_chart.set_categories(cats)

        std_chart.series[0].graphicalProperties.solidFill = "FF9800"

        ws.add_chart(std_chart, "J30")

    def _create_distribution_analysis(self, ws, responses):
        """Create distribution analysis section"""
        valid_probs = [r.probability for r in responses if r.probability is not None]

        if not valid_probs:
            return

        # Distribution statistics
        ws['A25'] = "📊 PROBABILITY DISTRIBUTION ANALYSIS"
        ws['A25'].style = self.header_style
        ws.merge_cells('A25:H25')

        # Calculate percentiles
        percentiles = [10, 25, 50, 75, 90]
        dist_stats = []

        for p in percentiles:
            value = np.percentile(valid_probs, p)
            dist_stats.append((f"{p}th Percentile", f"{value:.1f}%"))

        # Add quartile information
        q1 = np.percentile(valid_probs, 25)
        q3 = np.percentile(valid_probs, 75)
        iqr = q3 - q1
        dist_stats.extend([
            ("Interquartile Range (IQR)", f"{iqr:.1f}%"),
            ("Skewness", f"{self._calculate_skewness(valid_probs):.3f}"),
            ("Kurtosis", f"{self._calculate_kurtosis(valid_probs):.3f}")
        ])

        for i, (stat, value) in enumerate(dist_stats):
            ws[f'A{26+i}'] = stat
            ws[f'A{26+i}'].font = Font(bold=True)
            ws[f'B{26+i}'] = value
            ws[f'B{26+i}'].style = self.data_style

    def _calculate_skewness(self, data):
        """Calculate skewness of distribution"""
        from scipy.stats import skew
        return skew(data)

    def _calculate_kurtosis(self, data):
        """Calculate kurtosis of distribution"""
        from scipy.stats import kurtosis
        return kurtosis(data)

    def _create_visual_analysis_sheet(self, workbook: Workbook, export_data: ExportData, chart_dir: Path):
        """Create visual analysis sheet with embedded matplotlib charts"""
        ws = workbook.create_sheet("📈 Visual Analysis")

        # Title
        ws['A1'] = "📈 Visual Analysis Dashboard"
        ws['A1'].style = self.title_style
        ws.merge_cells('A1:H1')

        # Add explanation
        ws['A3'] = "This sheet contains advanced visualizations generated using matplotlib/seaborn"
        ws['A3'].font = Font(italic=True, size=11)
        ws.merge_cells('A3:H3')

        # Embed charts
        try:
            row_offset = 5

            # Model comparison chart
            comparison_path = chart_dir / "model_comparison.png"
            if comparison_path.exists():
                ws[f'A{row_offset}'] = "🤖 Model Performance Comparison"
                ws[f'A{row_offset}'].style = self.header_style
                ws.merge_cells(f'A{row_offset}:H{row_offset}')

                img = Image(str(comparison_path))
                img.width = 800
                img.height = 400
                ws.add_image(img, f'A{row_offset+1}')
                row_offset += 25

            # Distribution chart
            dist_path = chart_dir / "probability_distribution.png"
            if dist_path.exists():
                ws[f'A{row_offset}'] = "📊 Probability Distribution Analysis"
                ws[f'A{row_offset}'].style = self.header_style
                ws.merge_cells(f'A{row_offset}:H{row_offset}')

                img = Image(str(dist_path))
                img.width = 800
                img.height = 400
                ws.add_image(img, f'A{row_offset+1}')
                row_offset += 25

            # Timeline chart
            timeline_path = chart_dir / "forecast_timeline.png"
            if timeline_path.exists():
                ws[f'A{row_offset}'] = "⏰ Forecast Timeline"
                ws[f'A{row_offset}'].style = self.header_style
                ws.merge_cells(f'A{row_offset}:H{row_offset}')

                img = Image(str(timeline_path))
                img.width = 800
                img.height = 400
                ws.add_image(img, f'A{row_offset+1}')

        except Exception as e:
            logger.warning(f"Could not embed charts in visual analysis sheet: {e}")
            ws['A5'] = f"❌ Could not load visualizations: {e}"
            ws['A5'].font = Font(color="FF0000")

        # Overall statistics
        stats = export_data.forecast_result.statistics
        ws['A3'] = "Overall Statistics"
        ws['A3'].font = Font(size=14, bold=True)

        overall_stats = [
            ("Mean", f"{stats.mean:.2f}%" if stats.mean else "N/A"),
            ("Median", f"{stats.median:.2f}%" if stats.median else "N/A"),
            ("Standard Deviation", f"{stats.std:.2f}%" if stats.std else "N/A"),
            ("Minimum", f"{stats.min:.2f}%" if stats.min else "N/A"),
            ("Maximum", f"{stats.max:.2f}%" if stats.max else "N/A"),
            ("Range", f"{stats.max - stats.min:.2f}%" if stats.max and stats.min else "N/A")
        ]

        for i, (stat, value) in enumerate(overall_stats):
            row = 4 + i
            ws[f'A{row}'] = stat
            ws[f'B{row}'] = value

        # Model comparison table
        ws['A12'] = "Model Comparison"
        ws['A12'].font = Font(size=14, bold=True)

        # Convert model stats to DataFrame
        model_data = []
        for model, model_stats in stats.model_stats.items():
            model_data.append({
                'Model': model.split('/')[-1],
                'Mean': model_stats.mean,
                'Std_Dev': model_stats.std,
                'Min': model_stats.min,
                'Max': model_stats.max,
                'Count': model_stats.count,
                'Success_Rate': model_stats.success_rate * 100
            })

        df = pd.DataFrame(model_data)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Format the table
        start_row = 13
        for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=start_row + len(model_data))):
            for cell in row:
                if i == 0:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    def _create_reasoning_analysis_sheet(self, workbook: Workbook, export_data: ExportData):
        """Create reasoning analysis sheet"""
        ws = workbook.create_sheet("Reasoning Analysis")

        # Title
        ws['A1'] = "Reasoning Analysis"
        ws['A1'].font = Font(size=16, bold=True)

        # Key themes section
        ws['A3'] = "Common Reasoning Themes"
        ws['A3'].font = Font(size=14, bold=True)

        # Analyze reasoning content
        reasoning_texts = []
        for response in export_data.forecast_result.responses:
            if response.content and response.status.value == "success":
                reasoning_texts.append(response.content)

        # Extract common themes (simplified analysis)
        common_themes = self._extract_themes(reasoning_texts)

        for i, theme in enumerate(common_themes):
            ws[f'A{4+i}'] = f"• {theme}"

        # Model-specific patterns
        ws['A15'] = "Model-Specific Patterns"
        ws['A15'].font = Font(size=14, bold=True)

        # Group responses by model
        model_responses = {}
        for response in export_data.forecast_result.responses:
            if response.status.value == "success" and response.content:
                model = response.model.split('/')[-1]
                if model not in model_responses:
                    model_responses[model] = []
                model_responses[model].append(response.content)

        row = 16
        for model, responses in model_responses.items():
            # Avoid merged cells by using different columns
            ws.cell(row=row, column=1, value=f"{model}:")
            ws.cell(row=row, column=1).font = Font(bold=True)

            # Simple analysis of response characteristics
            avg_length = sum(len(r) for r in responses) / len(responses)
            has_base_rate = sum(1 for r in responses if 'base rate' in r.lower()) / len(responses)

            ws.cell(row=row, column=2, value=f"Avg length: {avg_length:.0f} chars, Base rate usage: {has_base_rate*100:.0f}%")
            row += 1

    def _create_metadata_sheet(self, workbook: Workbook, export_data: ExportData):
        """Create metadata sheet with configuration and prompt"""
        ws = workbook.create_sheet("Metadata")

        # Title
        ws['A1'] = "Metadata & Configuration"
        ws['A1'].font = Font(size=16, bold=True)

        # Forecast metadata
        metadata = export_data.forecast_result.metadata
        ws['A3'] = "Forecast Information"
        ws['A3'].font = Font(size=14, bold=True)

        metadata_items = [
            ("Question", metadata.question),
            ("Definition", metadata.definition),
            ("Timeframe", metadata.timeframe or "Not specified"),
            ("Models Used", ", ".join([m.split('/')[-1] for m in metadata.models])),
            ("Iterations per Model", metadata.iterations_per_model),
            ("Total Queries", metadata.total_queries),
            ("Start Time", metadata.start_time),
            ("End Time", metadata.end_time),
            ("Duration", f"{metadata.duration_seconds:.2f} seconds")
        ]

        for i, (key, value) in enumerate(metadata_items):
            row = 4 + i
            ws[f'A{row}'] = key
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = str(value)

        # Prompt used
        ws['A15'] = "Prompt Template Used"
        ws['A15'].font = Font(size=14, bold=True)
        ws['A16'] = export_data.forecast_result.prompt
        ws['A16'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[16].height = 200

        # Export information
        ws['A20'] = "Export Information"
        ws['A20'].font = Font(size=14, bold=True)
        ws['A21'] = "Export Timestamp"
        ws['A21'].font = Font(bold=True)
        ws['B21'] = export_data.export_timestamp
        ws['A22'] = "Tool Version"
        ws['A22'].font = Font(bold=True)
        ws['B22'] = "Foresight Analyzer v1.0"

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 80

    def _extract_themes(self, texts: List[str]) -> List[str]:
        """
        Extract common themes from reasoning texts

        Args:
            texts: List of reasoning texts

        Returns:
            List of common themes
        """
        # Simple keyword-based theme extraction
        themes = []

        # Count mentions of key concepts
        keywords = {
            "Base Rate Analysis": ["base rate", "historical", "reference class"],
            "Geopolitical Factors": ["geopolitical", "diplomatic", "international"],
            "Economic Considerations": ["economic", "financial", "cost"],
            "Military Factors": ["military", "war", "conflict", "army"],
            "Timeline Concerns": ["time", "timeline", "deadline", "duration"],
            "Uncertainty": ["uncertain", "unpredictable", "unknown"],
            "Stakeholder Analysis": ["stakeholder", "parties", "actors"],
            "External Pressure": ["pressure", "international", "sanctions"]
        }

        for theme, words in keywords.items():
            count = sum(1 for text in texts if any(word in text.lower() for word in words))
            if count > len(texts) * 0.3:  # If mentioned in >30% of responses
                themes.append(f"{theme} (mentioned in {count}/{len(texts)} responses)")

        return themes[:10]  # Return top 10 themes

    def _create_advanced_algorithms_sheet(self, workbook: Workbook, export_data: ExportData):
        """Create advanced algorithms analysis sheet showing enhanced ensemble results"""
        ws = workbook.create_sheet("🧠 Advanced Algorithms")

        # Title
        ws['A1'] = "🧠 Advanced Scientific Algorithms Analysis"
        ws['A1'].style = self.title_style
        ws.merge_cells('A1:H1')
        ws.row_dimensions[1].height = 30

        # Description
        ws['A3'] = "This sheet shows results from advanced forecasting algorithms that enhance basic ensemble averaging:"
        ws['A3'].font = Font(size=12, italic=True)
        ws.merge_cells('A3:H3')

        # Check if advanced analysis exists
        stats = export_data.forecast_result.statistics
        advanced_analysis = None

        # Try different possible attribute names for advanced analysis
        if hasattr(stats, 'advanced_analysis') and stats.advanced_analysis:
            advanced_analysis = stats.advanced_analysis
        elif hasattr(stats, '__dict__'):
            # Look for advanced analysis in the raw statistics
            for key, value in stats.__dict__.items():
                if 'advanced' in str(key).lower() and isinstance(value, dict):
                    advanced_analysis = value
                    break

        if not advanced_analysis:
            ws['A5'] = "⚠️ Advanced algorithms not available in this analysis"
            ws['A5'].font = Font(size=14, bold=True, color="FF6600")
            ws.merge_cells('A5:H5')

            # Debug info
            ws['A7'] = f"Available attributes: {list(stats.__dict__.keys()) if hasattr(stats, '__dict__') else 'No __dict__'}"
            ws['A7'].font = Font(size=10, italic=True, color="999999")
            ws.merge_cells('A7:H7')
            return

        advanced = advanced_analysis

        # Algorithm Results Section
        ws['A5'] = "🔬 ALGORITHM RESULTS"
        ws['A5'].style = self.header_style
        ws.merge_cells('A5:H5')

        # Basic vs Enhanced comparison
        basic_mean = stats.mean if hasattr(stats, 'mean') else 0
        bayesian_pred = advanced.get('bayesian_ensemble', basic_mean)
        calibrated_pred = advanced.get('calibrated_ensemble', None)

        results_data = [
            ("Basic Ensemble (Simple Mean)", f"{basic_mean:.1f}%", "Standard arithmetic average"),
            ("Bayesian Ensemble", f"{bayesian_pred:.1f}%", "Robust aggregation with outlier detection"),
            ("Calibrated Ensemble", f"{calibrated_pred:.1f}%" if calibrated_pred else "N/A", "Temperature-scaled probability calibration")
        ]

        # Headers
        headers = ["Method", "Result", "Description"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=7, column=i+1)
            cell.value = header
            cell.style = self.header_style

        # Data
        for i, (method, result, desc) in enumerate(results_data):
            row = 8 + i
            ws[f'A{row}'] = method
            ws[f'B{row}'] = result
            ws[f'C{row}'] = desc

            # Styling
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].style = self.data_style
            if "Bayesian" in method:
                ws[f'B{row}'].fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")

        # Model Consistency Scores
        ws['A12'] = "📊 MODEL CONSISTENCY SCORES"
        ws['A12'].style = self.header_style
        ws.merge_cells('A12:H12')

        consistency_scores = advanced.get('model_consistency_scores', {})
        if consistency_scores:
            # Headers
            ws['A14'] = "Model"
            ws['B14'] = "Consistency Score"
            ws['C14'] = "Quality Rating"
            for col in ['A14', 'B14', 'C14']:
                ws[col].style = self.header_style

            # Data
            for i, (model, score) in enumerate(consistency_scores.items()):
                row = 15 + i
                model_name = model.split('/')[-1] if '/' in model else model
                quality = "High" if score > 0.8 else "Medium" if score > 0.6 else "Low"

                ws[f'A{row}'] = model_name
                ws[f'B{row}'] = f"{score:.3f}"
                ws[f'C{row}'] = quality

                # Color coding for quality
                if quality == "High":
                    ws[f'C{row}'].fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                elif quality == "Medium":
                    ws[f'C{row}'].fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                else:
                    ws[f'C{row}'].fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

        # Quality Assessment
        quality_assessment = advanced.get('quality_assessment', {})
        if quality_assessment:
            ws['A20'] = "📈 QUALITY METRICS"
            ws['A20'].style = self.header_style
            ws.merge_cells('A20:H20')

            variance_reduction = quality_assessment.get('variance_reduction', 0)
            consensus_improvement = quality_assessment.get('consensus_improvement', {})

            metrics = [
                ("Variance Reduction", f"{variance_reduction:.1f}%", "How much algorithms reduced prediction variance"),
                ("Algorithm Status", "✅ Active" if variance_reduction > 0 else "⚠️ No Improvement", "Advanced algorithms effectiveness")
            ]

            for i, (metric, value, desc) in enumerate(metrics):
                row = 22 + i
                ws[f'A{row}'] = metric
                ws[f'B{row}'] = value
                ws[f'C{row}'] = desc

                ws[f'A{row}'].font = Font(bold=True)
                if "✅" in value:
                    ws[f'B{row}'].fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                elif "⚠️" in value:
                    ws[f'B{row}'].fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

        # Calibrated Predictions Detail
        calibrated_predictions = advanced.get('calibrated_predictions', {})
        if calibrated_predictions:
            ws['A27'] = "🎯 CALIBRATED PREDICTIONS BY MODEL"
            ws['A27'].style = self.header_style
            ws.merge_cells('A27:H27')

            # Headers
            ws['A29'] = "Model"
            ws['B29'] = "Raw Predictions"
            ws['C29'] = "Calibrated Predictions"
            for col in ['A29', 'B29', 'C29']:
                ws[col].style = self.header_style

            # Data
            for i, (model, cal_probs) in enumerate(calibrated_predictions.items()):
                row = 30 + i
                model_name = model.split('/')[-1] if '/' in model else model

                # Get original predictions for comparison
                model_stats = stats.model_stats.get(model, None)
                raw_mean = f"{model_stats.mean:.1f}%" if model_stats else "N/A"
                cal_mean = f"{np.mean(cal_probs):.1f}%" if cal_probs else "N/A"

                ws[f'A{row}'] = model_name
                ws[f'B{row}'] = raw_mean
                ws[f'C{row}'] = cal_mean

        # Auto-adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 20

    def _create_probability_heatmap_sheet(self, workbook: Workbook, export_data: ExportData):
        """Create probability distribution heatmap visualization"""
        ws = workbook.create_sheet("🔥 Probability Heatmap")

        # Title
        ws['A1'] = "📊 Probability Distribution Heatmap"
        ws['A1'].style = self.title_style
        ws.merge_cells('A1:J1')
        ws.row_dimensions[1].height = 30

        # Collect all probabilities by model
        model_probabilities = {}
        for response in export_data.forecast_result.responses:
            if response.probability is not None:
                if response.model not in model_probabilities:
                    model_probabilities[response.model] = []
                model_probabilities[response.model].append(response.probability)

        if not model_probabilities:
            ws['A3'] = "No probability data available"
            return

        # Create probability bins (0-10, 10-20, ..., 90-100)
        bins = list(range(0, 110, 10))
        bin_labels = [f"{i}-{i+10}%" for i in range(0, 100, 10)]

        # Headers
        ws['A3'] = "Model"
        ws['A3'].style = self.header_style
        for i, label in enumerate(bin_labels):
            cell = ws.cell(row=3, column=i+2)
            cell.value = label
            cell.style = self.header_style

        # Calculate distribution for each model
        row = 4
        for model, probs in model_probabilities.items():
            model_name = model.split('/')[-1] if '/' in model else model
            ws[f'A{row}'] = model_name
            ws[f'A{row}'].font = Font(bold=True)

            # Count probabilities in each bin
            hist, _ = np.histogram(probs, bins=bins)
            max_count = max(hist) if max(hist) > 0 else 1

            for i, count in enumerate(hist):
                cell = ws.cell(row=row, column=i+2)
                cell.value = count

                # Color intensity based on count (heatmap effect)
                if count > 0:
                    intensity = int(255 - (count / max_count) * 200)
                    color = f"{intensity:02X}FF{intensity:02X}"  # Green gradient
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

            row += 1

        # Add consensus indicator row
        row += 1
        ws[f'A{row}'] = "📊 CONSENSUS"
        ws[f'A{row}'].style = self.header_style

        # Calculate overall distribution
        all_probs = [p for probs in model_probabilities.values() for p in probs]
        overall_hist, _ = np.histogram(all_probs, bins=bins)
        max_overall = max(overall_hist) if max(overall_hist) > 0 else 1

        for i, count in enumerate(overall_hist):
            cell = ws.cell(row=row, column=i+2)
            cell.value = count

            # Highlight the bin with most predictions
            if count == max_overall and count > 0:
                cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
                cell.font = Font(bold=True)
            elif count > 0:
                intensity = int(255 - (count / max_overall) * 150)
                color = f"FF{intensity:02X}{intensity:02X}"  # Red gradient for consensus
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

        # Add summary statistics
        row += 3
        ws[f'A{row}'] = "📈 Distribution Insights"
        ws[f'A{row}'].style = self.header_style
        ws.merge_cells(f'A{row}:D{row}')

        # Find mode bin
        mode_bin_idx = np.argmax(overall_hist)
        mode_bin = bin_labels[mode_bin_idx]

        # Calculate concentration
        total_predictions = sum(overall_hist)
        mode_concentration = (overall_hist[mode_bin_idx] / total_predictions * 100) if total_predictions > 0 else 0

        insights = [
            ("Most Common Range", mode_bin, f"{mode_concentration:.1f}% of predictions"),
            ("Total Predictions", total_predictions, f"Across {len(model_probabilities)} models"),
            ("Consensus Strength", self._calculate_consensus_strength(overall_hist), "Based on distribution"),
        ]

        for i, (metric, value, description) in enumerate(insights):
            r = row + 1 + i
            ws[f'A{r}'] = metric
            ws[f'B{r}'] = value
            ws[f'C{r}'] = description
            ws[f'A{r}'].font = Font(bold=True)
            ws.merge_cells(f'C{r}:D{r}')

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws.column_dimensions[col].width = 12

    def _create_executive_brief_sheet(self, workbook: Workbook, export_data: ExportData):
        """Create auto-generated executive brief"""
        ws = workbook.create_sheet("📋 Executive Brief", 0)  # Insert as first sheet

        # Professional header
        ws['A1'] = "EXECUTIVE BRIEF"
        ws['A1'].font = Font(size=18, bold=True, color="2E5C8A")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 35

        # Date and metadata
        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws['A2'].font = Font(italic=True, size=10)
        ws.merge_cells('A2:F2')

        # Question
        ws['A4'] = "FORECAST QUESTION"
        ws['A4'].font = Font(bold=True, size=12, color="366092")
        ws.merge_cells('A4:F4')

        ws['A5'] = export_data.forecast_result.metadata.question
        ws['A5'].font = Font(size=11)
        ws['A5'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells('A5:F6')
        ws.row_dimensions[5].height = 40

        # Main probability result (large and prominent)
        ensemble_prob = export_data.forecast_result.get_ensemble_probability()
        ws['A8'] = "FORECAST PROBABILITY"
        ws['A8'].font = Font(bold=True, size=12, color="366092")
        ws.merge_cells('A8:C8')

        if ensemble_prob:
            ws['A9'] = f"{ensemble_prob:.1f}%"
            ws['A9'].font = Font(size=36, bold=True, color="E74C3C")
            ws['A9'].alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells('A9:C10')
            ws.row_dimensions[9].height = 50

            # Confidence interval
            std = export_data.forecast_result.statistics.std if export_data.forecast_result.statistics.std else 0
            lower = max(0, ensemble_prob - std)
            upper = min(100, ensemble_prob + std)
            ws['A11'] = f"Confidence Interval: {lower:.0f}% - {upper:.0f}%"
            ws['A11'].font = Font(italic=True, size=10)
            ws['A11'].alignment = Alignment(horizontal="center")
            ws.merge_cells('A11:C11')

        # Model consensus indicator
        ws['D8'] = "MODEL CONSENSUS"
        ws['D8'].font = Font(bold=True, size=12, color="366092")
        ws.merge_cells('D8:F8')

        consensus = self._get_consensus_level(export_data.forecast_result.statistics)
        consensus_color = "27AE60" if "Strong" in consensus else "F39C12" if "Moderate" in consensus else "E74C3C"
        ws['D9'] = consensus
        ws['D9'].font = Font(size=20, bold=True, color=consensus_color)
        ws['D9'].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('D9:F10')

        # Key factors section
        ws['A13'] = "KEY FORECAST DRIVERS"
        ws['A13'].font = Font(bold=True, size=12, color="366092")
        ws.merge_cells('A13:F13')

        # Extract key factors from responses
        key_factors = self._extract_key_factors(export_data.forecast_result.responses)
        for i, factor in enumerate(key_factors[:3], 1):
            row = 13 + i
            ws[f'A{row}'] = f"{i}."
            ws[f'B{row}'] = factor
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].alignment = Alignment(wrap_text=True)
            ws.merge_cells(f'B{row}:F{row}')

        # Model performance summary
        row = 18
        ws[f'A{row}'] = "MODEL PERFORMANCE SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="366092")
        ws.merge_cells(f'A{row}:F{row}')

        # Performance metrics
        row += 1
        metrics_headers = ["Metric", "Value"]
        for i, header in enumerate(metrics_headers):
            cell = ws.cell(row=row, column=i+1)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D5E8F7", end_color="D5E8F7", fill_type="solid")

        performance_metrics = [
            ("Models Used", len(export_data.forecast_result.metadata.models)),
            ("Success Rate", f"{export_data.forecast_result.statistics.success_rate * 100:.1f}%"),
            ("Average Response Quality", self._calculate_avg_quality_score(export_data.forecast_result)),
            ("Analysis Duration", f"{export_data.forecast_result.metadata.duration_seconds:.1f} seconds"),
        ]

        for metric, value in performance_metrics:
            row += 1
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(size=10)
            ws[f'B{row}'].font = Font(size=10)

        # Risk assessment
        row += 2
        ws[f'A{row}'] = "RISK ASSESSMENT"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="366092")
        ws.merge_cells(f'A{row}:F{row}')

        row += 1
        risk_level = self._assess_risk_level(ensemble_prob, export_data.forecast_result.statistics.std)
        risk_color = "27AE60" if risk_level == "Low" else "F39C12" if risk_level == "Medium" else "E74C3C"
        ws[f'A{row}'] = f"Risk Level: {risk_level}"
        ws[f'A{row}'].font = Font(size=11, bold=True, color=risk_color)
        ws.merge_cells(f'A{row}:F{row}')

        # Recommendations
        row += 2
        ws[f'A{row}'] = "RECOMMENDATIONS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="366092")
        ws.merge_cells(f'A{row}:F{row}')

        recommendations = self._generate_recommendations(ensemble_prob, consensus, risk_level)
        for i, rec in enumerate(recommendations, 1):
            row += 1
            ws[f'A{row}'] = f"• {rec}"
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            ws.merge_cells(f'A{row}:F{row}')

        # Footer
        row += 2
        ws[f'A{row}'] = "This brief was automatically generated based on ensemble AI forecasting analysis."
        ws[f'A{row}'].font = Font(italic=True, size=9, color="808080")
        ws[f'A{row}'].alignment = Alignment(horizontal="center")
        ws.merge_cells(f'A{row}:F{row}')

        # Format column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

        # Add borders to main sections
        for section_start in [4, 8, 13, 18]:
            for col in range(1, 7):
                ws.cell(row=section_start, column=col).border = Border(top=Side(style='medium'))

    def _calculate_consensus_strength(self, histogram):
        """Calculate consensus strength from probability histogram"""
        if sum(histogram) == 0:
            return "N/A"
        max_concentration = max(histogram) / sum(histogram)
        if max_concentration > 0.6:
            return "Very Strong"
        elif max_concentration > 0.4:
            return "Strong"
        elif max_concentration > 0.25:
            return "Moderate"
        else:
            return "Weak"

    def _extract_key_factors(self, responses):
        """Extract key factors from model responses"""
        factors = []
        keywords = ["military", "diplomatic", "economic", "political", "sanctions", "NATO", "weapons", "negotiations"]

        for response in responses[:5]:  # Sample first 5 responses
            if response.content:
                content_lower = response.content.lower()
                for keyword in keywords:
                    if keyword in content_lower and keyword not in [f.lower() for f in factors]:
                        # Extract sentence containing keyword
                        sentences = response.content.split('.')
                        for sent in sentences:
                            if keyword in sent.lower():
                                factor = sent.strip()[:100] + "..." if len(sent) > 100 else sent.strip()
                                if factor and factor not in factors:
                                    factors.append(factor)
                                break
                        if len(factors) >= 3:
                            break
            if len(factors) >= 3:
                break

        return factors if factors else ["Military developments", "Diplomatic negotiations", "Economic pressures"]

    def _calculate_avg_quality_score(self, forecast_result):
        """Calculate average quality score based on response characteristics"""
        total_score = 0
        count = 0

        for response in forecast_result.responses:
            if response.content:
                score = 0
                # Length score
                length = len(response.content)
                if length > 3000:
                    score += 3
                elif length > 1500:
                    score += 2
                elif length > 500:
                    score += 1

                # Probability extraction success
                if response.probability is not None:
                    score += 2

                total_score += score
                count += 1

        if count == 0:
            return "N/A"

        avg = total_score / count
        if avg >= 4:
            return "⭐⭐⭐⭐⭐ Excellent"
        elif avg >= 3:
            return "⭐⭐⭐⭐ Good"
        elif avg >= 2:
            return "⭐⭐⭐ Moderate"
        else:
            return "⭐⭐ Basic"

    def _assess_risk_level(self, probability, std):
        """Assess risk level based on probability and uncertainty"""
        if std is None:
            std = 0

        if probability > 70:
            if std < 10:
                return "High"
            else:
                return "Medium-High"
        elif probability > 40:
            if std < 15:
                return "Medium"
            else:
                return "Medium-Uncertain"
        else:
            if std < 10:
                return "Low"
            else:
                return "Low-Uncertain"

    def _generate_recommendations(self, probability, consensus, risk_level):
        """Generate recommendations based on analysis results"""
        recommendations = []

        if probability > 70:
            recommendations.append("High probability event - prepare for likely scenario occurrence")
        elif probability > 40:
            recommendations.append("Moderate probability - maintain flexible planning for multiple scenarios")
        else:
            recommendations.append("Low probability - monitor developments but focus on alternative scenarios")

        if "Strong" in consensus:
            recommendations.append("Strong model agreement increases confidence in forecast")
        elif "Weak" in consensus:
            recommendations.append("Weak consensus suggests high uncertainty - seek additional analysis")

        if "Uncertain" in risk_level:
            recommendations.append("High variance in predictions - consider expanding analysis scope")

        return recommendations

    def _calculate_model_confidence(self, stats, responses, model):
        """Calculate enhanced confidence score for a model"""
        confidence_score = {"score": 0, "stars": 1, "factors": {}}

        # Factor 1: Consistency (lower std dev = higher confidence)
        consistency_score = 0
        if stats.std > 0:
            consistency_score = max(0, 1 - (stats.std / 30))  # Normalize std to 0-1 scale
        else:
            consistency_score = 1
        confidence_score["factors"]["consistency"] = consistency_score * 0.3

        # Factor 2: Success rate
        success_score = stats.success_rate
        confidence_score["factors"]["success_rate"] = success_score * 0.3

        # Factor 3: Response quality (length and probability extraction)
        quality_scores = []
        model_responses = [r for r in responses if r.model == model]
        for response in model_responses[:5]:  # Sample first 5 responses
            if response.content:
                length_score = min(1, len(response.content) / 3000)
                prob_score = 1 if response.probability is not None else 0
                quality_scores.append((length_score + prob_score) / 2)

        avg_quality = sum(quality_scores) / len(quality_scores) if quality_scores else 0
        confidence_score["factors"]["response_quality"] = avg_quality * 0.25

        # Factor 4: Sample size reliability
        sample_score = min(1, stats.count / 10)  # More samples = higher confidence
        confidence_score["factors"]["sample_size"] = sample_score * 0.15

        # Calculate total score
        total_score = sum(confidence_score["factors"].values())
        confidence_score["score"] = total_score

        # Convert to star rating
        if total_score >= 0.85:
            confidence_score["stars"] = 5
        elif total_score >= 0.7:
            confidence_score["stars"] = 4
        elif total_score >= 0.5:
            confidence_score["stars"] = 3
        elif total_score >= 0.3:
            confidence_score["stars"] = 2
        else:
            confidence_score["stars"] = 1

        return confidence_score